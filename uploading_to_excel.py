import openpyxl as xl
from pathlib import Path


def upload_to_excel(list_of_dictionaries,tags_according_to_file):
        
    # Finding and opening excel file, wb, and sheet.
    filepath = Path(r"C:\Users\brennan.ukrainetz\Desktop\Coding\Python\cal_sheet_project\transactions.xlsx")
    wb = xl.load_workbook(filepath)
    obj_sheet = wb["Sheet1"]

    index = 0
    for dictionary in list_of_dictionaries:       

        # Deciding which column the data is stored in.
        error_string = ""
        for key, value in dictionary.items():
            if key == "TAG":
                col = 1
                if value != tags_according_to_file[index]:
                    error_string = "_DIFFTAG_" + tags_according_to_file[index]  #If the value doesn't match the tag, the tag will have ERROR appended to it.
            elif key == "CALIBRATED RANGE":
                col = 2
            elif key == "MAKE":
                col = 3
            elif key == "MODEL":
                col = 4
            elif key == "SERIAL":
                col = 5
            elif key == "SIZE":
                col = 6
            elif key == "DEV ID":
                col = 7
            elif key == "BLURB":
                col = 8
            else:
                raise ValueError("KEY NOT FOUND.")
            # After column is picked, put the value in it.

            working_cell = obj_sheet.cell(row = index+2, column = col)
            working_cell.value = value + error_string
            # print("Cell's value is: " + str(working_cell.value))
            error_string = "" #Reset in event that error_string activated.
        
        
        index += 1

    wb.save(filepath) #Save file.


# tags_according_to_file = "PT-2685"
# dictionaries = [{'BLURB': 'LIN BACKUP PUMPS', 'TAG': 'PT-2685', 'CALIBRATED RANGE': '0-6000 KPA', 'MAKE': 'ROSEMOUNT', 'MODEL': '', 'SERIAL': ''}, {'BLURB': 'LIasdfasd PUMPS', 'TAG': 'PT-422685', 'CALIBRATED RANGE': 'asdf00 KPA', 'MAKE': 'ROdddNT', 'MODEL': 'das', 'SERIAL': ''}]

# upload_to_excel(dictionaries,tags_according_to_file)
















# INSTRUCTIONAL CODE- LEARNING WITH MOSH
# print(wb.sheetnames)
# cell = obj_sheet["A1"]
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
# # Another way to use cell.
# cell = obj_sheet.cell(row=1,column=1)
# # obj_sheet.max_column
# # obj_sheet.max_row


# Note: sheetnames is an attribute.
# values = [1,2,3]
# for row in range(1,obj_sheet.max_row+1):

# for column in range(1,len(values)): # range(1,obj_sheet.max_column+1):
#     working_cell = obj_sheet.cell(1,column)
#     working_cell.value = values[1]
#     print(working_cell.value)
